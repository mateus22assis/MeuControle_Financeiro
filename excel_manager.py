from openpyxl import load_workbook
from datetime import datetime
from calendar import monthrange

# ==========================
# CONFIGURAÇÕES
# ==========================

CAMINHO_PLANILHA = "ControleFinanceiro_2026_Prototipo_v3.xlsx"

# ==========================
# UTILITÁRIOS
# ==========================

def abrirPlanilha():
    """
    Abre a planilha principal do controle financeiro.

    Utiliza o caminho configurado para o arquivo Excel do projeto.

    Retorna o objeto de workbook carregado.
    """

    return load_workbook(CAMINHO_PLANILHA)


def encontrarPrimeiraLinhaVazia(
    aba,
    coluna="A",
    linhaInicial=2
):
    """
    Localiza a primeira linha sem valor em uma coluna da aba.

    A busca começa na linha inicial informada para preservar cabeçalhos.

    Retorna o número da primeira linha vazia encontrada.
    """

    linha = linhaInicial

    while aba[f"{coluna}{linha}"].value is not None:
        linha += 1

    return linha

def ordenarMovimentacoesPorData():
    """
    Ordena os registros da aba Movimentacoes pela data.

    Preserva o cabeçalho ao regravar as movimentações ordenadas.

    Não retorna valor.
    Ordena a aba Movimentacoes pela coluna Data.
    """

    workbook = abrirPlanilha()

    aba = workbook["Movimentacoes"]

    movimentacoes = list(
        aba.iter_rows(
            min_row=2,
            values_only=True
        )
    )

    movimentacoes.sort(
        key=lambda linha: converterData(
            linha[0]
        )
    )

    aba.delete_rows(
        2,
        aba.max_row
    )

    for movimentacao in movimentacoes:
        aba.append(movimentacao)

    workbook.save(CAMINHO_PLANILHA)

def converterData(data):
    """
    Converte uma data para o formato datetime aceito pelo sistema.

    Aceita datas já convertidas e os formatos de texto previstos.

    Retorna a data convertida ou gera erro para formato inválido.
    """

    if isinstance(data, datetime):
        return data

    formatos = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(data, formato)
        except ValueError:
            pass

    raise ValueError(
        f"Formato de data inválido: {data}"
    )


# ==========================
# LEITURA DA PLANILHA
# ==========================

def lerConfiguracoes():
    """
    Lê as configurações financeiras armazenadas na planilha.

    Mantém valores padrão quando algum campo não possui valor informado.

    Retorna um dicionário com as configurações disponíveis.
    """

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    configuracoes = {
        "receitaMensal": 0.0,
        "percentualReserva": 30.0,
        "limiteCartao": 0.0,
        "diaFechamento": 3,
        "diaVencimento": 10
    }

    for linha in aba_configuracoes.iter_rows(
            min_row=2,
            values_only=True
    ):

        campo, valor = linha

        if campo is None:
            continue

        campo = campo.strip().lower()

        if campo == "receita mensal":

            configuracoes["receitaMensal"] = (
                valor if valor is not None else 0.0
            )

        elif campo == "percentual reserva":

            configuracoes["percentualReserva"] = (
                valor if valor is not None else 30.0
            )

        elif campo == "limite cartao":

            configuracoes["limiteCartao"] = (
                valor if valor is not None else 0.0
            )

        elif campo == "dia fechamento":

            configuracoes["diaFechamento"] = (
                valor if valor is not None else 0
            )

        elif campo == "dia vencimento":

            configuracoes["diaVencimento"] = (
                valor if valor is not None else 0
            )

    return configuracoes


def lerCompromissosMensais():
    """
    Lê os compromissos mensais cadastrados na planilha.

    Considera apenas linhas que possuem uma descrição informada.

    Retorna uma lista de compromissos com descrição e valor.
    """

    workbook = abrirPlanilha()

    aba_compromissos = workbook["CompromissosMensais"]

    compromissos = []

    for linha in aba_compromissos.iter_rows(
            min_row=2,
            values_only=True
    ):

        descricao = linha[0]
        valor = linha[1]

        if descricao is not None:

            compromissos.append({

                "descricao": descricao,
                "valor": valor

            })

    return compromissos


def lerMovimentacoes():
    """
    Lê as movimentações financeiras cadastradas na planilha.

    Considera apenas linhas com descrição e preserva o número original
    da linha para permitir consultas e exclusões.

    Retorna uma lista de movimentações registradas.
    """

    workbook = abrirPlanilha()

    aba_movimentacoes = workbook["Movimentacoes"]

    movimentacoes = []

    for numeroLinha, linha in enumerate(
        aba_movimentacoes.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):
        
        data = linha[0]
        natureza = linha[1]
        meio = linha[2]
        categoria = linha[3]
        descricao = linha[4]
        valor = linha[5]
        parcelas = linha[6]

        if descricao is not None:

            movimentacoes.append({
                "linha": numeroLinha,
                "data": data,
                "natureza": natureza,
                "meio": meio,
                "categoria": categoria,
                "descricao": descricao,
                "valor": valor,
                "parcelas": parcelas

            })

    return movimentacoes


# ==========================
# ESCRITA DA PLANILHA
# ==========================

def salvarReceitaMensal(valor):
    """
    Salva a receita mensal informada nas configurações.

    Atualiza o campo correspondente na aba Configuracoes.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    aba_configuracoes["B2"] = valor

    workbook.save(CAMINHO_PLANILHA)


def salvarPercentualReserva(valor):
    """
    Salva o percentual destinado à reserva financeira.

    Atualiza o campo correspondente na aba Configuracoes.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    aba_configuracoes["B3"] = valor

    workbook.save(CAMINHO_PLANILHA)


def limparFaturas():
    """
    Remove os registros de faturas já gravados na planilha.

    Preserva a primeira linha da aba Faturas, destinada ao cabeçalho.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_faturas = workbook["Faturas"]

    if aba_faturas.max_row > 1:
        aba_faturas.delete_rows(
            2,
            aba_faturas.max_row - 1
        )

    workbook.save(CAMINHO_PLANILHA)


def gravarResumoFaturas(resumoFaturas):
    """
    Grava na planilha o resumo de faturas recebido.

    Limpa os registros anteriores antes de inserir os novos dados.

    Não retorna valor.
    """

    limparFaturas()

    workbook = abrirPlanilha()

    aba_faturas = workbook["Faturas"]

    for fatura in resumoFaturas:
        aba_faturas.append([
            fatura["fatura"],
            fatura["vencimento"],
            fatura["valor"],
            fatura["status"]
        ])

    workbook.save(CAMINHO_PLANILHA)


# ==========================
# ATUALIZAÇÃO DE ABAS
# ==========================

def atualizarPlanilha():
    """
    Atualiza as abas derivadas da planilha financeira.

    Neste momento, gera e grava somente o resumo da aba Faturas.
    Novas abas derivadas poderão ser incluídas neste ponto de atualização.

    Não retorna valor.
    """

    from calculos import gerarResumoFaturas

    resumoFaturas = gerarResumoFaturas()

    gravarResumoFaturas(resumoFaturas)


def adicionarMovimentacao(
    natureza,
    meio,
    categoria,
    descricao,
    valor,
    parcelas="",
    dataMovimentacao=None
):
    """
    Adiciona uma movimentação financeira à planilha.

    Quando a data não é informada, utiliza a data atual e ordena
    as movimentações após salvar o novo registro.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_movimentacoes = workbook["Movimentacoes"]

    proximaLinha = encontrarPrimeiraLinhaVazia(
        aba_movimentacoes
    )

    if dataMovimentacao is None:
        dataMovimentacao = datetime.now().strftime("%d/%m/%Y")

    aba_movimentacoes[f"A{proximaLinha}"] = dataMovimentacao
    aba_movimentacoes[f"B{proximaLinha}"] = natureza
    aba_movimentacoes[f"C{proximaLinha}"] = meio
    aba_movimentacoes[f"D{proximaLinha}"] = categoria
    aba_movimentacoes[f"E{proximaLinha}"] = descricao
    aba_movimentacoes[f"F{proximaLinha}"] = valor
    aba_movimentacoes[f"G{proximaLinha}"] = ""

    workbook.save(CAMINHO_PLANILHA)

    ordenarMovimentacoesPorData()


def adicionarCompromissoMensal(
    descricao,
    valor
):
    """
    Adiciona um compromisso mensal à planilha.

    Insere a descrição e o valor na próxima linha disponível da aba.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_compromissos = workbook["CompromissosMensais"]

    proximaLinha = encontrarPrimeiraLinhaVazia(
        aba_compromissos
    )

    aba_compromissos[f"A{proximaLinha}"] = descricao
    aba_compromissos[f"B{proximaLinha}"] = valor

    workbook.save(CAMINHO_PLANILHA)

def excluirMovimentacao(linha):
    """
    Exclui uma movimentação pelo número de sua linha na planilha.

    Não permite remover cabeçalhos, linhas inexistentes ou linhas vazias.

    Retorna True quando a exclusão é realizada e False caso contrário.
    """

    workbook = abrirPlanilha()

    aba_movimentacoes = workbook["Movimentacoes"]

    if linha < 2 or linha > aba_movimentacoes.max_row:
        return False
    if aba_movimentacoes.cell(row=linha, column=1).value is None:
        return False
    aba_movimentacoes.delete_rows(linha)

    workbook.save(CAMINHO_PLANILHA)

    return True

# ==========================
# REGRAS DO CARTÃO
# ==========================

def calcularMesFatura(dataCompra):
    """
    Determina a fatura correspondente a uma compra no cartão.

    Compras após o fechamento são atribuídas ao mês seguinte.

    Recebe uma data (string no formato dd/mm/AAAA)
    e retorna o mês/ano da fatura correspondente.
    """

    configuracoes = lerConfiguracoes()

    diaFechamento = configuracoes["diaFechamento"]

    data = datetime.strptime(
        dataCompra,
        "%d/%m/%Y"
    )

    mes = data.month
    ano = data.year

    if data.day > diaFechamento:

        mes += 1

        if mes > 12:
            mes = 1
            ano += 1

    return f"{mes:02d}/{ano}"


def adicionarMeses(data, quantidadeMeses):
    """
    Adiciona meses a uma data utilizada pelo sistema.

    Ajusta o dia para o último dia do mês de destino quando necessário.

    Recebe uma data (string dd/mm/AAAA)
    e retorna outra data adicionando meses.
    """

    data = datetime.strptime(
        data,
        "%d/%m/%Y"
    )

    mes = data.month + quantidadeMeses
    ano = data.year

    while mes > 12:
        mes -= 12
        ano += 1

    ultimoDia = monthrange(
        ano,
        mes
    )[1]

    dia = min(
        data.day,
        ultimoDia
    )

    novaData = datetime(
        ano,
        mes,
        dia
    )

    return novaData.strftime("%d/%m/%Y")
