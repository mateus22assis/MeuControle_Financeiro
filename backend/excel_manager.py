from openpyxl import load_workbook
from datetime import datetime
from calendar import monthrange
import re
import os
import sys


if getattr(sys, "frozen", False):
    PASTA_PROGRAMA = os.path.dirname(sys.executable)
else:
    PASTA_PROGRAMA = os.path.dirname(
        os.path.dirname(os.path.abspath(__file__))
    )


CAMINHO_PLANILHA = os.path.join(
    PASTA_PROGRAMA,
    "ControleFinanceiro_2026.xlsx"
)


print("CAMINHO DA PLANILHA:", CAMINHO_PLANILHA)
print("PLANILHA EXISTE:", os.path.exists(CAMINHO_PLANILHA))
# ==========================
# UTILITÁRIOS
# ==========================

def abrirPlanilha():
    """
    Abre a planilha principal do controle financeiro.

    Utiliza o caminho configurado para o arquivo Excel do projeto.

    Retorna o objeto de workbook carregado.
    """

    return load_workbook(CAMINHO_PLANILHA)


def encontrarPrimeiraLinhaVazia(
    aba,
    coluna="A",
    linhaInicial=2
):
    """
    Localiza a primeira linha sem valor em uma coluna da aba.

    A busca começa na linha inicial informada para preservar cabeçalhos.

    Retorna o número da primeira linha vazia encontrada.
    """

    linha = linhaInicial

    while aba[f"{coluna}{linha}"].value is not None:
        linha += 1

    return linha

def ordenarMovimentacoesPorData():
    """
    Ordena os registros da aba Movimentacoes pela data.

    Preserva o cabeçalho ao regravar as movimentações ordenadas.

    Não retorna valor.
    Ordena a aba Movimentacoes pela coluna Data.
    """

    workbook = abrirPlanilha()

    aba = workbook["Movimentacoes"]

    movimentacoes = list(
        aba.iter_rows(
            min_row=2,
            values_only=True
        )
    )

    movimentacoes.sort(
        key=lambda linha: (
            linha[0] is None or str(linha[0]).strip() == "",
            converterData(linha[0]) if linha[0] is not None else datetime.max
        )
    )

    if aba.max_row > 1:
        aba.delete_rows(2, aba.max_row - 1)

    for movimentacao in movimentacoes:
        aba.append(movimentacao)

    workbook.save(CAMINHO_PLANILHA)
    # ... (Refatoração para usar as novas funções centralizadas)

def converterData(data):
    """
    Converte uma data para o formato datetime aceito pelo sistema.

    Aceita datas já convertidas e os formatos de texto previstos.

    Retorna a data convertida ou None para dados vazios.
    Gera erro para formato inválido de texto.
    """

    if isinstance(data, datetime):
        return data

    if data is None or str(data).strip() == "":
        return None

    formatos = [
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
    ]

    for formato in formatos:
        try:
            return datetime.strptime(str(data).strip(), formato)
        except (ValueError, TypeError):
            pass

    raise ValueError(
        f"Formato de data inválido: {data}"
    )



# ==========================
# LEITURA DA PLANILHA
# ==========================

def lerConfiguracoes():
    """
    Lê as configurações financeiras armazenadas na planilha.

    Mantém valores padrão quando algum campo não possui valor informado.

    Retorna um dicionário com as configurações disponíveis.
    """

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    configuracoes = {
        "receitaMensal": 0.0,
        "percentualReserva": 30.0,
        "saldoInicial": 0.0,
        "limiteCartao": 0.0,
        "diaFechamento": 3,
        "diaVencimento": 10
    }

    for linha in aba_configuracoes.iter_rows(
            min_row=2,
            values_only=True
    ):

        campo, valor = linha

        if campo is None:
            continue

        campo = campo.strip().lower()

        if campo == "receita mensal":

            configuracoes["receitaMensal"] = (
                valor if valor is not None else 0.0
            )

        elif campo == "saldo inicial":

            configuracoes["saldoInicial"] = (
                valor if valor is not None else 0.0
            )

        elif campo == "percentual reserva":


            configuracoes["percentualReserva"] = (
                valor if valor is not None else 30.0
            )

        elif campo == "limite cartao":

            configuracoes["limiteCartao"] = (
                valor if valor is not None else 0.0
            )

        elif campo == "dia fechamento":

            configuracoes["diaFechamento"] = (
                valor if valor is not None else 0
            )

        elif campo == "dia vencimento":

            configuracoes["diaVencimento"] = (
                valor if valor is not None else 0
            )

    return configuracoes


def lerCompromissosMensais():
    """
    Lê os compromissos mensais cadastrados na planilha.

    Considera apenas linhas que possuem uma descrição informada.

    Retorna uma lista de compromissos com descrição e valor.
    """

    workbook = abrirPlanilha()

    aba_compromissos = workbook["CompromissosMensais"]

    compromissos = []

    for numeroLinha, linha in enumerate (aba_compromissos.iter_rows(
            min_row=2,
            values_only=True),
            start=2
    ):

        descricao = linha[0]
        valor = linha[1]

        if descricao is not None:

            compromissos.append({
                "linha": numeroLinha,
                "descricao": descricao,
                "valor": valor
                

            })

    return compromissos


def lerMovimentacoes():
    """
    Lê as movimentações financeiras cadastradas na planilha.

    Considera apenas linhas com descrição e preserva o número original
    da linha para permitir consultas e exclusões.

    Retorna uma lista de movimentações registradas.
    """

    workbook = abrirPlanilha()

    aba_movimentacoes = workbook["Movimentacoes"]

    movimentacoes = []

    for numeroLinha, linha in enumerate(
        aba_movimentacoes.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):
        
        data = linha[0]
        natureza = linha[1]
        meio = linha[2]
        categoria = linha[3]
        descricao = linha[4]
        valor = linha[5]
        parcelas = linha[6]

        if descricao is not None:

            movimentacoes.append({
                "linha": numeroLinha,
                "data": data,
                "natureza": natureza,
                "meio": meio,
                "categoria": categoria,
                "descricao": descricao,
                "valor": valor,
                "parcelas": parcelas

            })

    return movimentacoes

def lerCategorias():
  '''
   Lê as categorias cadastradas na planilha.

    Considera o nome, a natureza e o status de cada categoria.

    Retorna uma lista de categorias.
    '''  
  workbook = abrirPlanilha()

  aba_categorias = workbook["Categorias"]

  categorias = []

  for linha in aba_categorias.iter_rows(
          min_row=2,    
        values_only=True
    ):

      nome = linha[0]
      natureza = linha[1]
      ativa = linha[2]

      if nome is not None:
            categorias.append({
                "nome": nome,
                "natureza": natureza,
                "ativa": str(ativa).lower() == "sim"
            })

  return categorias


# ==========================
# ESCRITA DA PLANILHA
# ==========================

def salvarReceitaMensal(valor):
    """
    Salva a receita mensal informada nas configurações.

    Atualiza o campo correspondente na aba Configuracoes.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    aba_configuracoes["B2"] = valor

    workbook.save(CAMINHO_PLANILHA)


def salvarPercentualReserva(valor):
    """
    Salva o percentual destinado à reserva financeira.

    Atualiza o campo correspondente na aba Configuracoes.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    aba_configuracoes["B3"] = valor

    workbook.save(CAMINHO_PLANILHA)

def salvarLimiteCartao(valor):

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    aba_configuracoes["B4"] = valor

    workbook.save(CAMINHO_PLANILHA)


def salvarDiaFechamento(valor):

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    aba_configuracoes["B5"] = valor

    workbook.save(CAMINHO_PLANILHA)


def salvarDiaVencimento(valor):

    workbook = abrirPlanilha()

    aba_configuracoes = workbook["Configuracoes"]

    aba_configuracoes["B6"] = valor

    workbook.save(CAMINHO_PLANILHA)

def limparFaturas():
    """
    Remove os registros de faturas já gravados na planilha.

    Preserva a primeira linha da aba Faturas, destinada ao cabeçalho.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_faturas = workbook["Faturas"]

    if aba_faturas.max_row > 1:
        aba_faturas.delete_rows(
            2,
            aba_faturas.max_row - 1
        )

    workbook.save(CAMINHO_PLANILHA)


# ==========================
# RESUMO DE FATURAS
# ==========================
def lerAbatimentosFaturas():
    """
    Lê os valores abatidos previamente de cada fatura.

    Retorna um dicionário no formato:
    {
        "09/2026": 100.0
    }
    """

    workbook = abrirPlanilha()

    aba_faturas = workbook["Faturas"]

    abatimentos = {}

    for linha in aba_faturas.iter_rows(
        min_row=2,
        values_only=True
    ):
        fatura = linha[0]

        if fatura is None:
            continue

        valorAbatido = (
            linha[3]
            if len(linha) > 3 and linha[3] is not None
            else 0.0
        )

        abatimentos[str(fatura)] = float(valorAbatido)

    return abatimentos


def gravarResumoFaturas(resumoFaturas):
    """
    Grava o resumo das faturas preservando
    os valores abatidos previamente.
    """

    abatimentos = lerAbatimentosFaturas()

    limparFaturas()

    workbook = abrirPlanilha()

    aba_faturas = workbook["Faturas"]

    for fatura in resumoFaturas:

        valorAbatidoPreviamente = abatimentos.get(
            fatura["fatura"],
            0.0
        )

        aba_faturas.append([
            fatura["fatura"],
            fatura["vencimento"],
            fatura["valor"],
            valorAbatidoPreviamente,
            fatura["status"]
        ])

    workbook.save(CAMINHO_PLANILHA)

def registrarAbatimentoFatura(valor):
    """
    Registra um valor abatido previamente
    na fatura atualmente aberta.

    Retorna True quando o abatimento é registrado
    e False quando o valor é inválido.
    """

    if valor is None or float(valor) <= 0:
        return False

    workbook = abrirPlanilha()

    aba_faturas = workbook["Faturas"]

    valor = float(valor)

    for linha in range(2, aba_faturas.max_row + 1):

        status = aba_faturas[f"E{linha}"].value

        if status != "Aberta":
            continue

        valorPrevisto = float(
            aba_faturas[f"C{linha}"].value or 0
        )

        valorAbatido = float(
            aba_faturas[f"D{linha}"].value or 0
        )

        valorEmAberto = (
            valorPrevisto - valorAbatido
        )

        if valor > valorEmAberto:
            return False

        aba_faturas[f"D{linha}"] = (
            valorAbatido + valor
        )

        workbook.save(CAMINHO_PLANILHA)

        return True

    return False   

# ==========================
# ATUALIZAÇÃO DE ABAS
# ==========================

def atualizarPlanilha():
    """
    Atualiza as abas derivadas da planilha financeira.

    Neste momento, gera e grava somente o resumo da aba Faturas.
    Novas abas derivadas poderão ser incluídas neste ponto de atualização.

    Não retorna valor.
    """
    from backend.calculos import gerarResumoFaturas
    resumoFaturas = gerarResumoFaturas()

    gravarResumoFaturas(resumoFaturas)

def adicionarMovimentacao(
    natureza,
    meio,
    categoria,
    descricao,
    valor,
    parcelas="",
    dataMovimentacao=None
):
    """
    Adiciona uma movimentação financeira à planilha.

    Quando a data não é informada, utiliza a data atual e ordena
    as movimentações após salvar o novo registro.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_movimentacoes = workbook["Movimentacoes"]

    proximaLinha = encontrarPrimeiraLinhaVazia(
        aba_movimentacoes
    )

    if dataMovimentacao is None:
        dataMovimentacao = datetime.now().strftime("%d/%m/%Y")

    aba_movimentacoes[f"A{proximaLinha}"] = dataMovimentacao
    aba_movimentacoes[f"B{proximaLinha}"] = natureza
    aba_movimentacoes[f"C{proximaLinha}"] = meio
    aba_movimentacoes[f"D{proximaLinha}"] = categoria
    aba_movimentacoes[f"E{proximaLinha}"] = descricao
    aba_movimentacoes[f"F{proximaLinha}"] = valor
    aba_movimentacoes[f"G{proximaLinha}"] = parcelas

    workbook.save(CAMINHO_PLANILHA)

    ordenarMovimentacoesPorData()


def adicionarCompromissoMensal(
    descricao,
    valor
):
    """
    Adiciona um compromisso mensal à planilha.

    Insere a descrição e o valor na próxima linha disponível da aba.

    Não retorna valor.
    """

    workbook = abrirPlanilha()

    aba_compromissos = workbook["CompromissosMensais"]

    proximaLinha = encontrarPrimeiraLinhaVazia(
        aba_compromissos
    )

    aba_compromissos[f"A{proximaLinha}"] = descricao
    aba_compromissos[f"B{proximaLinha}"] = valor

    workbook.save(CAMINHO_PLANILHA)
def alterarCompromissoMensal(linha, descricao, valor):
    """
    Altera um compromisso mensal existente.

    Recebe o número da linha na planilha, a nova descrição
    e o novo valor.

    Retorna True quando a alteração é realizada e False
    quando a linha é inválida ou está vazia.
    """

    workbook = abrirPlanilha()

    aba_compromissos = workbook["CompromissosMensais"]

    if linha < 2 or linha > aba_compromissos.max_row:
        return False

    if aba_compromissos.cell(row=linha, column=1).value is None:
        return False

    aba_compromissos.cell(
        row=linha,
        column=1
    ).value = descricao

    aba_compromissos.cell(
        row=linha,
        column=2
    ).value = valor

    workbook.save(CAMINHO_PLANILHA)

    return True


def excluirCompromissoMensal(linha):
    """
    Exclui um compromisso mensal pelo número de sua linha
    na planilha.

    Retorna True quando a exclusão é realizada e False
    quando a linha é inválida ou está vazia.
    """

    workbook = abrirPlanilha()

    aba_compromissos = workbook["CompromissosMensais"]

    if linha < 2 or linha > aba_compromissos.max_row:
        return False

    if aba_compromissos.cell(row=linha, column=1).value is None:
        return False

    aba_compromissos.delete_rows(linha)

    workbook.save(CAMINHO_PLANILHA)

    return True

def excluirMovimentacao(linha):
    """
    Exclui uma movimentação pelo número de sua linha na planilha.

    Não permite remover cabeçalhos, linhas inexistentes ou linhas vazias.

    Retorna True quando a exclusão é realizada e False caso contrário.
    """

    workbook = abrirPlanilha()

    aba_movimentacoes = workbook["Movimentacoes"]

    if linha < 2 or linha > aba_movimentacoes.max_row:
        return 0
    if aba_movimentacoes.cell(row=linha, column=1).value is None:
        return 0

    movimentacaoSelecionada = list(
        aba_movimentacoes.iter_rows(
            min_row=linha,
            max_row=linha,
            values_only=True
        )
    )[0]

    parcelasSelecionadas = re.fullmatch(
        r"(.+) \(([1-9]\d*)/([1-9]\d*)\)",
        str(movimentacaoSelecionada[4])
    )

    linhasParaExcluir = [linha]

    if parcelasSelecionadas is not None:
        descricaoBase = parcelasSelecionadas.group(1)
        totalParcelas = int(parcelasSelecionadas.group(3))
        natureza = movimentacaoSelecionada[1]
        meio = movimentacaoSelecionada[2]
        categoria = movimentacaoSelecionada[3]
        valor = movimentacaoSelecionada[5]
        linhasParaExcluir = []

        for numeroLinha, movimentacao in enumerate(
            aba_movimentacoes.iter_rows(
                min_row=2,
                values_only=True
            ),
            start=2
        ):
            parcelasMovimentacao = re.fullmatch(
                r"(.+) \(([1-9]\d*)/([1-9]\d*)\)",
                str(movimentacao[4])
            )

            if parcelasMovimentacao is None:
                continue

            numeroParcela = int(parcelasMovimentacao.group(2))
            totalParcelasMovimentacao = int(parcelasMovimentacao.group(3))

            if (
                parcelasMovimentacao.group(1) == descricaoBase
                and 1 <= numeroParcela <= totalParcelas
                and totalParcelasMovimentacao == totalParcelas
                and movimentacao[1] == natureza
                and movimentacao[2] == meio
                and movimentacao[3] == categoria
                and movimentacao[5] == valor
            ):
                linhasParaExcluir.append(numeroLinha)

    for numeroLinha in sorted(linhasParaExcluir, reverse=True):
        aba_movimentacoes.delete_rows(numeroLinha)

    workbook.save(CAMINHO_PLANILHA)

    atualizarPlanilha()

    return len(linhasParaExcluir)



def anteciparParcelas(linhas, dataAntecipacao=None):
   """ Altera a data das parcelas selecionadas para a data da antecipação. Recebe uma lista com os números das linhas das parcelas que serão antecipadas. Quando a data não é informada, utiliza a data atual. Retorna a quantidade de parcelas alteradas.
   """

   workbook = abrirPlanilha()

   aba_movimentacoes = workbook["Movimentacoes"]

   if dataAntecipacao is None:
       dataAntecipacao = datetime.now().strftime("%d/%m/%Y")

   dataAntecipacao = converterData(dataAntecipacao)

   if dataAntecipacao is None:
       return 0

   parcelasAlteradas = 0

   for linha in linhas:

       if linha < 2 or linha > aba_movimentacoes.max_row:
           continue

       movimentacao = aba_movimentacoes.cell(row=linha, column=1).value

       if movimentacao is None:
           continue

       descricao = aba_movimentacoes.cell(row=linha, column=5).value

       parcelas = re.fullmatch(
           r"(.+) \(([1-9]\d*)/([1-9]\d*)\)", str(descricao) )

       if parcelas is None:
           continue

       aba_movimentacoes.cell(row=linha, column=1).value = dataAntecipacao

       parcelasAlteradas += 1

   workbook.save(CAMINHO_PLANILHA)

   ordenarMovimentacoesPorData()

   atualizarPlanilha()

   return parcelasAlteradas
def adicionarCategoria(nome, natureza):

    workbook = abrirPlanilha()

    aba_categorias = workbook["Categorias"]

    proximaLinha = encontrarPrimeiraLinhaVazia(
        aba_categorias
    )

    aba_categorias[f"A{proximaLinha}"] = nome
    aba_categorias[f"B{proximaLinha}"] = natureza
    aba_categorias[f"C{proximaLinha}"] = "Sim"


    workbook.save(CAMINHO_PLANILHA)
def alterarStatusCategoria(nome, natureza, ativa):

    workbook = abrirPlanilha()

    aba_categorias = workbook["Categorias"]

    for linha in range(2, aba_categorias.max_row + 1):

        nomeCategoria = aba_categorias[f"A{linha}"].value
        naturezaCategoria = aba_categorias[f"B{linha}"].value

        if (
            nomeCategoria == nome
            and naturezaCategoria == natureza
        ):
            aba_categorias[f"C{linha}"] = (
                "Sim" if ativa else "Não"
            )

            break

    workbook.save(CAMINHO_PLANILHA)  
# ==========================
# REGRAS DO CARTÃO
# ==========================

def determinarMesFatura(dataCompra, diaFechamento):
    """
    Determina o mês e ano da fatura para uma compra.
    Retorna uma tupla (mês, ano).
    """
    data = converterData(dataCompra)
    if data is None:
        return None, None

    mes = data.month
    ano = data.year

    if data.day >= diaFechamento:
        mes += 1
        if mes > 12:
            mes = 1
            ano += 1
    
    return mes, ano

def calcularMesFatura(dataCompra):
    """
    Determina a fatura correspondente a uma compra no cartão.

    Compras após o fechamento são atribuídas ao mês seguinte.

    Recebe uma data (string ou datetime)
    e retorna o mês/ano da fatura correspondente em texto MM/AAAA.
    """

    configuracoes = lerConfiguracoes()
    diaFechamento = configuracoes["diaFechamento"]

    mes, ano = determinarMesFatura(dataCompra, diaFechamento)
    
    if mes is None:
        return None

    return f"{mes:02d}/{ano}"



def adicionarMeses(data, quantidadeMeses):
    """
    Adiciona meses a uma data utilizada pelo sistema.

    Ajusta o dia para o último dia do mês de destino quando necessário.

    Recebe uma data (string dd/mm/AAAA)
    e retorna outra data adicionando meses.
    """

    data = datetime.strptime(
        data,
        "%d/%m/%Y"
    )

    mes = data.month + quantidadeMeses
    ano = data.year

    while mes > 12:
        mes -= 12
        ano += 1

    ultimoDia = monthrange(
        ano,
        mes
    )[1]

    dia = min(
        data.day,
        ultimoDia
    )

    novaData = datetime(
        ano,
        mes,
        dia
    )

    return novaData.strftime("%d/%m/%Y")
